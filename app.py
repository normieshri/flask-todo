from flask import Flask,render_template, redirect,request, url_for
from flask_sqlalchemy import SQLAlchemy
from flask import send_file
from openpyxl import load_workbook
from datetime import date,datetime

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///site.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = True
db = SQLAlchemy(app)

class Todo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title =  db.Column(db.String(100))
    disc = db.Column(db.String(300))
    completed = db.Column(db.Boolean)


@app.route("/hello")
def hello_world():
    return render_template("base.html")

@app.route("/")
def todo():
    todo = Todo.query.all()
    return render_template("todo.html",todo = todo)

@app.route("/add", methods=["POST"])
def create_todo():
    title = request.form.get("title")
    disc = request.form.get("disc")
    if title == "":
        title = "You did not give any title :("
    new_todo = Todo(title=title,disc = disc,completed=False)
    db.session.add(new_todo)
    db.session.commit()
    data = (Todo.query.filter_by(title = title).all())[-1]
    info = {"action":"CREATE","id":data.id,"title":data.title,"desc":data.disc,"status":data.completed}
    write_report(info)
    return redirect(url_for("todo"))

@app.route("/update/<int:todo_id>")
def update(todo_id):
    todo = Todo.query.filter_by(id=todo_id).first()
    todo.completed = not todo.completed
    db.session.commit()
    data = Todo.query.filter_by(id = todo_id).first()
    info = {"action":"UPDATE","id":data.id,"title":data.title,"desc":data.disc,"status":data.completed}
    write_report(info)
    return redirect(url_for("todo"))


@app.route("/delete/<int:todo_id>")
def delete(todo_id):
    todo = Todo.query.filter_by(id=todo_id).first()
    data = Todo.query.filter_by(id = todo_id).first()
    info = {"action":"DELETE","id":data.id,"title":data.title,"desc":data.disc,"status":data.completed}
    write_report(info)
    db.session.delete(todo)
    db.session.commit()
    return redirect(url_for("todo"))
@app.route("/download")
def download_report():
    path = "./report.xlsx"
    return send_file(path, as_attachment=True)

def write_report(info):
    wb = load_workbook("report.xlsx")
    ws = wb["report"]
    today = date.today()
    now = datetime.now()
    cell = ws.max_row+1
    ws[f"A{cell}"] = now.strftime("%H:%M:%S")
    ws[f"B{cell}"] = today.strftime("%b-%d-%Y")
    ws[f"C{cell}"] = info["action"]
    ws[f"D{cell}"] = info["id"]
    ws[f"E{cell}"] = info["title"]
    ws[f"F{cell}"] = info["desc"]
    ws[f"G{cell}"] = info["status"]
    wb.save("report.xlsx")
    



if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        app.run(debug=True)
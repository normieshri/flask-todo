from openpyxl import load_workbook, Workbook
from datetime import date,datetime


wb = load_workbook("report.xlsx")
ws = wb["report"]
today = date.today()
now = datetime.now()
ws["A3"] = now.strftime("%H:%M:%S")
ws["B3"] = today.strftime("%b-%d-%Y")

print(ws.max_row)
wb.save("report.xlsx")